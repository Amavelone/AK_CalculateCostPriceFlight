from __future__ import annotations

import tempfile
import unittest
from datetime import time
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from app.modules.cost_monitor.catalog import tariffs_for_view
from app.modules.cost_monitor.parsers.monitor import parse_monitor_workbook
from app.modules.cost_monitor.records import CostMonitorDataset, MonitorWorkbookData
from app.modules.cost_monitor.source_adapters import (
    MonitorWorkbookSourceData,
    compatibility_adapter_for_parser,
    production_adapter_for_parser,
)
from app.modules.cost_monitor.sources import (
    fetch_usd_rate,
    mark_source_error,
    parse_fuel_registry,
    parse_srv_tariffs,
    refresh_source,
    save_uploaded_file,
    source_by_id,
    workbook_preview,
)
from openpyxl import Workbook


class SourceParserCharacterizationTests(unittest.TestCase):
    def test_srv_parser_keeps_max_kerosene_rate_and_source_order(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "srv.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["АП", "Услуга", "В/С", "Ставка", "Ед.изм.", "Дата с", "Дата по", "Наименование Орг"])
            worksheet.append(["KJA", "ВОДА", "738", 100, "рейс", "2026-01-01", "2026-12-31", "Тест"])
            worksheet.append(["KJA", "КЕРОСИН", "737", 80, "т", "2026-01-01", "2026-12-31", "Тест"])
            worksheet.append(["KJA", "КЕРОСИН", "738", 90, "т", "2026-02-01", "2026-12-31", "Тест"])
            worksheet.append(["AER", "ВОДА", "738", 150, "рейс", "2026-01-01", "2026-12-31", "Тест"])
            worksheet.append(["AER", "ВОДА", "738", 120, "рейс", "2026-01-01", "2026-12-31", "Тест"])
            worksheet.append(["KJA", "НЕИЗВЕСТНАЯ УСЛУГА", "738", 999, "рейс", None, None, "Тест"])
            workbook.save(path)

            tariffs, rows_read, preview, note = parse_srv_tariffs(path)
            adapter_run = production_adapter_for_parser("srv_tariffs").load(path)

        self.assertEqual(rows_read, 6)
        self.assertEqual([(row["service"], row["rate"]) for row in tariffs], [("ВОДА", 100.0), ("КЕРОСИН", 90.0), ("ВОДА", 150.0)])
        self.assertEqual([(row.service, row.rate) for row in adapter_run.data.tariffs], [("ВОДА", 100.0), ("КЕРОСИН", 90.0), ("ВОДА", 150.0)])
        self.assertEqual(len(preview), 6)
        self.assertIsNone(note)

    def test_fuel_parser_converts_usd_and_keeps_max_price_per_airport(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "fuel.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Служебная строка"])
            worksheet.append(["Партнер", "Валюта", "Цена", "Вид цены поставщика", "Период"])
            worksheet.append(["Поставщик 1", "USD", 1.5, "-|KJA|НДС сверху|USD", "2026-08"])
            worksheet.append(["Поставщик 2", "RUB", 140, "-|KJA|НДС сверху|RUB", "2026-08"])
            workbook.save(path)

            with patch("app.modules.cost_monitor.parsers.fuel.fetch_usd_rate", return_value=(100.0, "тестовый курс")):
                prices, rows_read, preview, note = parse_fuel_registry(path)
                adapter_run = production_adapter_for_parser("fuel_registry").load(path)

        self.assertEqual(rows_read, 2)
        self.assertEqual(prices[0]["airport"], "KJA")
        self.assertEqual(prices[0]["price"], 150.0)
        self.assertEqual(len(preview), 2)
        self.assertEqual(note, "тестовый курс")
        self.assertEqual(adapter_run.data.prices[0].price, 150.0)
        self.assertEqual(adapter_run.note, "тестовый курс")
        state = SourceParserCharacterizationTests.empty_dataset_state()
        adapter_run.data.apply(CostMonitorDataset.from_state(state)).write_to_state(state)
        self.assertEqual(state["fuel_prices"][0]["currency"], "USD")
        self.assertEqual(state["fuel_prices"][0]["exchange_rate"], 100.0)
        self.assertEqual(state["fuel_prices"][0]["exchange_rate_source"], "тестовый курс")
        self.assertFalse(state["fuel_prices"][0]["exchange_rate_fallback_used"])
        self.assertTrue(state["fuel_prices"][0]["exchange_rate_timestamp"])

    def test_monitor_parser_reads_all_configuration_sections(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "monitor.xlsx"
            workbook = Workbook()
            routes = workbook.active
            routes.title = "ИШР"
            routes.append(["№", "Откуда", "Куда", "Расстояние", "Время"])
            routes.append([1, "KJA", "OVB", 650, time(1, 30)])

            international = workbook.create_sheet("Признак МВЛ")
            international.append(["АП", "Признак"])
            international.append(["OVB", 1])

            directories = workbook.create_sheet("Справочники")
            directories.append([None] * 16)
            directory_row = [None] * 16
            directory_row[5:7] = ["738", 1.25]
            directory_row[11:16] = ["Базовый", "738", 1, 2, 3]
            directories.append(directory_row)

            other = workbook.create_sheet("Прочее")
            other.cell(1, 2, "OVB")
            other.cell(27, 2, 1234.5)

            manual = workbook.create_sheet("ЦРТ+")
            manual.append(["АП", "Услуга", "Ставка", "Ед.изм.", "В/С", "Дата с", "Дата по", "Наименование Орг"])
            manual.append(["OVB", "ВОДА", 500, "рейс", "738", "2026-01-01", "2026-12-31", "Тест"])
            workbook.save(path)

            result, rows_read, preview, note = parse_monitor_workbook(path)
            adapter_run = compatibility_adapter_for_parser("monitor_workbook").load(path)

        self.assertEqual(rows_read, 1)
        self.assertEqual(result["routes"][0]["flight_time"], 1.5)
        self.assertTrue(result["international_airports"]["OVB"])
        self.assertEqual(result["aircraft_multipliers"]["738"], 1.25)
        self.assertEqual(result["scenario_rates"]["Базовый"]["738"], [1.0, 2.0, 3.0])
        self.assertEqual(result["other_costs"]["OVB"], 1234.5)
        self.assertEqual(result["legacy_manual_tariffs"][0]["rate"], 500.0)
        self.assertEqual(len(preview), 1)
        self.assertIsNone(note)
        self.assertEqual(adapter_run.data.workbook.routes[0].key, "KJA-OVB")
        self.assertEqual(adapter_run.data.workbook.scenario_rates["Базовый"]["738"], (1.0, 2.0, 3.0))
        self.assertEqual(adapter_run.data.workbook.legacy_manual_tariffs[0].rate, 500.0)
        state = SourceParserCharacterizationTests.empty_dataset_state()
        state["manual_tariffs"] = [{"id": "manual-1", "airport": "KJA", "service": "ВОДА", "rate": 100, "source": "manual"}]
        adapter_run.data.apply(CostMonitorDataset.from_state(state)).write_to_state(state)
        self.assertEqual([item["id"] for item in state["manual_tariffs"]], ["manual-1", "legacy-2"])
        self.assertEqual(state["manual_tariffs"][1]["legacy_manual"], True)

    @staticmethod
    def empty_dataset_state() -> dict:
        return {
            "imported_tariffs": [],
            "manual_tariffs": [],
            "fuel_prices": [],
            "routes": [],
            "other_costs": {},
        }


class WorkbookPreviewTests(unittest.TestCase):
    def test_uses_detected_header_and_requested_worksheet(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "registry.xlsx"
            workbook = Workbook()
            workbook.active.title = "Титул"
            registry = workbook.create_sheet("Реестр")
            registry.append(["Служебная строка"])
            registry.append(["Дата выгрузки", "2026-08-28"])
            registry.append(["Партнер", "Валюта", "Цена"])
            registry.append(["Тестовый поставщик", "RUB", 100])
            registry.append(["Другой поставщик", "USD", 120])
            workbook.save(path)

            preview = workbook_preview(path, sheet_name="Реестр", row_limit=2)

            self.assertEqual(preview["sheet"], "Реестр")
            self.assertEqual(preview["sheets"], ["Титул", "Реестр"])
            self.assertEqual(preview["preview"][0]["Партнер"], "Тестовый поставщик")
            self.assertEqual(preview["preview"][1]["Цена"], 120)

    def test_reports_unknown_worksheet(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "book.xlsx"
            Workbook().save(path)

            with self.assertRaisesRegex(ValueError, "Лист не найден"):
                workbook_preview(path, sheet_name="Нет такого листа")

    def test_manual_service_is_retained_and_marked_as_conflict(self) -> None:
        state = {
            "imported_tariffs": [{"id": "file-1", "airport": "AAA", "service": "ВОДА", "rate": 100, "source": "file"}],
            "manual_tariffs": [{"id": "manual-1", "airport": "AAA", "service": "ВОДА", "rate": 200, "source": "manual"}],
        }

        tariffs = tariffs_for_view(state)

        self.assertFalse(tariffs[0]["conflict"])
        self.assertTrue(tariffs[1]["conflict"])
        self.assertEqual(tariffs[1]["source"], "manual")

    def test_missing_source_can_be_marked_with_actionable_error(self) -> None:
        state = {
            "source_configs": [
                {
                    "id": "srv",
                    "directory": "Z:/missing-source-directory",
                    "mask": "7480_srv*.xlsx",
                    "parser": "srv_tariffs",
                    "last_status": "ready",
                }
            ]
        }

        with self.assertRaises(FileNotFoundError):
            refresh_source(state, "srv", "2026-08-28T10:00:00+00:00")
        source = mark_source_error(state, "srv", "Директория не найдена", "2026-08-28T10:00:00+00:00")

        self.assertEqual(source["last_status"], "error")
        self.assertEqual(source["last_error"], "Директория не найдена")

    def test_cbr_failure_keeps_documented_current_fallback(self) -> None:
        with patch("app.modules.cost_monitor.parsers.fuel.httpx.get", side_effect=OSError("offline")):
            metadata = fetch_usd_rate()

        self.assertEqual(metadata.rate, 95.0)
        self.assertEqual(metadata.source, "fallback")
        self.assertTrue(metadata.fallback_used)
        self.assertTrue(metadata.timestamp)
        self.assertIn("резервное значение 95 RUB/USD", metadata.note)

    def test_compatibility_adapter_can_apply_empty_workbook_sections(self) -> None:
        state = {
            "manual_tariffs": [],
            "routes": [{"key": "OLD-OLD"}],
            "other_costs": {"OLD": 1},
        }
        MonitorWorkbookSourceData(MonitorWorkbookData.from_mapping({})).apply(
            CostMonitorDataset.from_state(state)
        ).write_to_state(state)

        self.assertEqual(state["routes"], [{"key": "OLD-OLD"}])

    def test_srv_adapter_normalizes_and_round_trips_canonical_data(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "srv.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["АП", "Услуга", "В/С", "Ставка", "Ед.изм.", "Дата с", "Дата по", "Наименование Орг"])
            worksheet.append(["KJA", "КЕРОСИН", "738", 90, "т", "2026-02-01", "2026-12-31", "Тест"])
            workbook.save(path)

            run = production_adapter_for_parser("srv_tariffs").load(path)

        self.assertEqual(run.source_id, "srv")
        self.assertEqual(run.data.rows_loaded, 1)
        state = SourceParserCharacterizationTests.empty_dataset_state()
        run.data.apply(CostMonitorDataset.from_state(state)).write_to_state(state)
        self.assertEqual(state["imported_tariffs"][0]["airport"], "KJA")
        self.assertEqual(state["imported_tariffs"][0]["rate"], 90.0)

    def test_production_adapter_lookup_rejects_compatibility_workbook(self) -> None:
        with self.assertRaisesRegex(ValueError, "production source adapter"):
            production_adapter_for_parser("monitor_workbook")

    def test_runtime_source_lookup_rejects_compatibility_workbook(self) -> None:
        with self.assertRaises(KeyError):
            source_by_id({"source_configs": [{"id": "monitor_workbook"}]}, "monitor_workbook")

    def test_invalid_upload_is_not_published(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            source = {"directory": temporary_directory}
            with self.assertRaises(Exception):
                save_uploaded_file(source, "invalid.xlsx", BytesIO(b"not an xlsx workbook"))
            self.assertFalse((Path(temporary_directory) / "invalid.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
