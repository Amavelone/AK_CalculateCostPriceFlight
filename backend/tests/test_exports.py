from __future__ import annotations

import json
import unittest
from io import BytesIO

from openpyxl import load_workbook

from app.modules.cost_monitor.calculation import calculate
from app.modules.cost_monitor.exports import build_export_snapshot, json_bytes, xlsx_bytes
from app.modules.cost_monitor.schemas import CalculationRequest


class ExportTests(unittest.TestCase):
    def state(self) -> dict:
        return {
            "routes": [{"key": "AAA-DME", "flight_time": 2, "distance": 100}],
            "international_airports": {},
            "imported_tariffs": [
                {"airport": "AAA", "service": "КЕРОСИН", "rate": 100},
                {"airport": "AAA", "service": "ЗАПРАВКА ВС", "rate": 20},
                {"airport": "AAA", "service": "АНО АД", "rate": 1000},
            ],
            "manual_tariffs": [],
            "fuel_prices": [],
            "scenario_rates": {"ГБ 2026": {"738": [10, 20, 30]}},
            "aircraft_multipliers": {"738": 1},
        }

    def request(self) -> CalculationRequest:
        return CalculationRequest.model_validate(
            {
                "legs": [{"id": "one", "departure": "AAA", "arrival": "DME", "aircraft": "738", "passengers": 17}],
                "settings": {"scenario": "ГБ 2026", "fuel_source": "ЦРТ", "catering": True},
            }
        )

    def test_snapshot_keeps_inputs_results_and_all_component_details(self) -> None:
        request = self.request()
        result = calculate(self.state(), request)
        snapshot = build_export_snapshot(request, result)

        self.assertEqual(snapshot["schema_version"], "1.0")
        self.assertEqual(snapshot["calculation"]["configuration"]["scenario"], "ГБ 2026")
        self.assertEqual(snapshot["calculation"]["legs"][0]["input"]["passengers"], 17)
        self.assertEqual(snapshot["calculation"]["totals"]["m2"], result["total"]["m2"])
        self.assertEqual(
            set(snapshot["calculation"]["legs"][0]["result"]["details"]),
            {"fuel", "ground", "ano", "catering", "vat"},
        )
        self.assertTrue(snapshot["calculation"]["legs"][0]["result"]["details"]["ano"])
        self.assertTrue(snapshot["calculation"]["legs"][0]["result"]["details"]["catering"])
        self.assertTrue(snapshot["calculation"]["legs"][0]["result"]["details"]["vat"])

    def test_json_and_xlsx_package_the_same_snapshot(self) -> None:
        request = self.request()
        result = calculate(self.state(), request)
        snapshot = build_export_snapshot(request, result)

        loaded_json = json.loads(json_bytes(snapshot).decode("utf-8"))
        self.assertEqual(loaded_json["calculation"]["totals"], snapshot["calculation"]["totals"])

        workbook = load_workbook(BytesIO(xlsx_bytes(snapshot)), data_only=True)
        self.assertEqual(workbook.sheetnames, ["РАСЧЕТ", "ДЕТАЛИЗАЦИЯ", "ПАРАМЕТРЫ"])
        calculation_sheet = workbook["РАСЧЕТ"]
        self.assertEqual(calculation_sheet["A1"].value, "Номер плеча")
        self.assertEqual(calculation_sheet["C2"].value, "AAA-DME")
        self.assertEqual(calculation_sheet["V2"].value, 17)
        self.assertEqual(calculation_sheet["O3"].value, snapshot["calculation"]["totals"]["m1"])
        self.assertEqual(calculation_sheet["P3"].value, snapshot["calculation"]["totals"]["m2"])
        self.assertEqual(calculation_sheet["Q3"].value, snapshot["calculation"]["totals"]["m3"])
        self.assertGreater(workbook["ДЕТАЛИЗАЦИЯ"].max_row, 1)


if __name__ == "__main__":
    unittest.main()
