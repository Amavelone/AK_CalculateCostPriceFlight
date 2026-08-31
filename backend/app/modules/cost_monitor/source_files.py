from __future__ import annotations

import uuid
from itertools import chain
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .catalog import normalize_key, normalize_text
from .parsers.common import normalize_for_json

MAX_UPLOAD_BYTES = 25 * 1024 * 1024


def find_latest_file(source: dict[str, Any]) -> Path:
    directory = Path(source["directory"])
    if not directory.is_dir():
        raise FileNotFoundError(f"Директория не найдена: {directory}")
    candidates = [path for path in directory.glob(source["mask"]) if path.is_file()]
    if not candidates:
        raise FileNotFoundError(f"Не найден файл по маске {source['mask']}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def find_active_file(source: dict[str, Any]) -> Path:
    """Возвращает файл уже активированного набора, не просто последний upload."""

    active_name = source.get("active_file")
    if not active_name:
        return find_latest_file(source)
    directory = Path(source["directory"])
    path = directory / Path(str(active_name)).name
    if not path.is_file():
        raise FileNotFoundError(f"Активный файл источника не найден: {path.name}")
    return path


def workbook_preview(path: Path, sheet_name: str | None = None, row_limit: int = 12) -> dict[str, Any]:
    """Возвращает читаемое превью выбранного листа Excel.

    Реестры могут содержать служебные строки перед таблицей, поэтому функция
    ищет заголовок в первых 25 строках. Интерфейс также может явно запросить
    любой видимый лист по имени.
    """

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name and sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Лист не найден: {sheet_name}")
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    leading_rows: list[tuple[Any, ...]] = []
    for _ in range(25):
        row = next(iterator, None)
        if row is None:
            break
        leading_rows.append(row)

    header_position = 0
    header_hints = {"АП", "АЭРОПОРТ", "ПАРТНЕР", "УСЛУГА", "ВИД ЦЕНЫ ПОСТАВЩИКА"}
    for index, row in enumerate(leading_rows):
        values = {normalize_key(value) for value in row if normalize_text(value)}
        if values.intersection(header_hints):
            header_position = index
            break
    header = leading_rows[header_position] if leading_rows else ()
    labels = [normalize_text(value) or f"Колонка {index + 1}" for index, value in enumerate(header)]
    preview: list[dict[str, Any]] = []
    remaining_rows = leading_rows[header_position + 1 :]
    for row in chain(remaining_rows, iterator):
        if len(preview) >= row_limit:
            break
        preview.append({labels[index]: normalize_for_json(value) for index, value in enumerate(row) if index < len(labels)})
    active_sheet = worksheet.title
    sheet_names = list(workbook.sheetnames)
    workbook.close()
    return {"sheet": active_sheet, "sheets": sheet_names, "preview": preview}


def save_uploaded_file(source: dict[str, Any], original_name: str, source_file: Any) -> Path:
    """Проверяет и атомарно публикует загруженный workbook в директории источника.

    Финальное имя файла не заменяется, пока ``openpyxl`` не примет временный
    файл. Поэтому неуспешная загрузка не может стать кандидатом на refresh.
    """

    file_name = Path(original_name or "source.xlsx").name
    if Path(file_name).suffix.lower() != ".xlsx":
        raise ValueError("Поддерживаются только файлы .xlsx")
    directory = Path(source["directory"])
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / file_name
    temporary = directory / f".{file_name}.{uuid.uuid4().hex}.uploading"
    written = 0
    try:
        with temporary.open("wb") as target_file:
            while chunk := source_file.read(1024 * 1024):
                written += len(chunk)
                if written > MAX_UPLOAD_BYTES:
                    raise ValueError(f"Размер файла превышает лимит {MAX_UPLOAD_BYTES // 1024 // 1024} МБ")
                target_file.write(chunk)
        # Проверяем фактическую структуру книги до публикации финального имени.
        workbook = load_workbook(temporary, read_only=True, data_only=True)
        workbook.close()
        temporary.replace(target)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return target
