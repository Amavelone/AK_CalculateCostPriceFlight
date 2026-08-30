from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..catalog import normalize_key, normalize_text
from .common import as_float, date_rank, header_index, value_by_key


ALLOWED_SERVICES = {
    "АНО АД",
    "АЭРОВОКЗАЛ",
    "АЭРОВОКЗАЛ(М)",
    "БУКСИРОВКА",
    "ВЗЛЕТ-ПОСАДКА",
    "ВОДА",
    "ЗАПРАВКА ВС",
    "КЕРОСИН",
    "ОБЕСПЕЧЕНИЕ МАСЛАМИ ТН",
    "ПАССАЖИР",
    "ПАССАЖИР(М)",
    "ПРИЕМ-ВЫПУСК",
    "САНУЗЕЛ",
    "СЛИВ ВОДЫ",
    "ТЕЛЕТРАП МИН",
    "ТРАНСПБЕЗОП",
    "ТРАНСПОРТ",
    "ТРАП",
    "УБОРКА",
    "БОРТПИТАНИЕ",
}

# Точные правила отбора AER из действующего определения Power Query ЦРТ_Check.
AER_RULES: dict[str, str | None] = {
    "БУКСИРОВКА": None,
    "ВОДА": "738",
    "ПРИЕМ-ВЫПУСК": "738",
    "САНУЗЕЛ": "738",
    "БОРТПИТАНИЕ": "738",
    "УБОРКА": "737",
}


def parse_srv_tariffs(path: Path) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]], str | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    header = next(iterator, ())
    index = header_index(header)
    rows_read = 0
    candidates: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []

    for row_number, row in enumerate(iterator, start=2):
        rows_read += 1
        airport = normalize_key(value_by_key(row, index, "АП"))
        service = normalize_key(value_by_key(row, index, "Услуга"))
        aircraft = normalize_key(value_by_key(row, index, "В/С"))
        rate = as_float(value_by_key(row, index, "Ставка"))

        if len(preview) < 12:
            preview.append({"АП": airport, "Услуга": service, "Ставка": rate, "В/С": aircraft})

        if not airport or service not in ALLOWED_SERVICES or rate is None:
            continue
        if aircraft not in {"", "733", "737", "738"}:
            continue
        candidates.append(
            {
                "id": f"srv-{row_number}",
                "airport": airport,
                "service": service,
                "rate": round(rate, 6),
                "unit": normalize_text(value_by_key(row, index, "Ед.изм.")),
                "aircraft": aircraft,
                "start_date": date_rank(value_by_key(row, index, "Дата с")),
                "end_date": date_rank(value_by_key(row, index, "Дата по")),
                "organization": normalize_text(value_by_key(row, index, "Наименование Орг")),
                "source": "file",
                "source_file": path.name,
                "source_row": row_number,
            }
        )
    workbook.close()

    # Действующий M-код оставляет один тариф керосина на аэропорт. При равной
    # ставке детерминированный порядок повторяет выбор по типу ВС и дате.
    kerosene_by_airport: dict[str, dict[str, Any]] = {}
    for tariff in candidates:
        if tariff["service"] != "КЕРОСИН":
            continue
        current = kerosene_by_airport.get(tariff["airport"])
        comparison_key = (tariff["rate"], tariff["aircraft"], tariff["start_date"])
        if current is None or comparison_key > (current["rate"], current["aircraft"], current["start_date"]):
            kerosene_by_airport[tariff["airport"]] = tariff

    tariffs = [
        tariff
        for tariff in candidates
        if tariff["service"] != "КЕРОСИН" or kerosene_by_airport[tariff["airport"]]["id"] == tariff["id"]
    ]

    # Power Query оставляет максимальные ставки AER для шести услуг; для пяти
    # из них дополнительно требуется конкретный тип воздушного судна.
    aer_max_rates: dict[str, float] = {}
    for tariff in tariffs:
        if tariff["airport"] == "AER" and tariff["service"] in AER_RULES:
            aer_max_rates[tariff["service"]] = max(aer_max_rates.get(tariff["service"], tariff["rate"]), tariff["rate"])

    tariffs = [
        tariff
        for tariff in tariffs
        if tariff["airport"] != "AER"
        or tariff["service"] not in AER_RULES
        or (
            tariff["rate"] == aer_max_rates.get(tariff["service"])
            and (AER_RULES[tariff["service"]] is None or tariff["aircraft"] == AER_RULES[tariff["service"]])
        )
    ]
    return tariffs, rows_read, preview, None
