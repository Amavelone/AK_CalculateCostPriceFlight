from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..catalog import normalize_key, normalize_text
from .common import as_float, as_hours, date_rank, header_index, value_by_key


def parse_monitor_workbook(path: Path) -> tuple[dict[str, Any], int, list[dict[str, Any]], str | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    routes: list[dict[str, Any]] = []
    international_airports: dict[str, bool] = {}
    other_costs: dict[str, float] = {}
    aircraft_multipliers: dict[str, float] = {}
    scenario_rates: dict[str, dict[str, list[float]]] = {}
    legacy_manual_tariffs: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []
    rows_read = 0

    if "ИШР" in workbook.sheetnames:
        worksheet = workbook["ИШР"]
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            rows_read += 1
            departure = normalize_key(row[1] if len(row) > 1 else None)
            arrival = normalize_key(row[2] if len(row) > 2 else None)
            distance = as_float(row[3] if len(row) > 3 else None) or 0.0
            flight_time = as_hours(row[4] if len(row) > 4 else None)
            if not departure or not arrival:
                continue
            route = {
                "key": f"{departure}-{arrival}",
                "departure": departure,
                "arrival": arrival,
                "distance": round(distance, 3),
                # Power Query сохраняет полное дробное время Excel. Не округляем
                # его до точности интерфейса: секунды влияют на топливо и маршрут.
                "flight_time": round(flight_time, 12),
                "source_row": row_number,
            }
            routes.append(route)
            if len(preview) < 12:
                preview.append(route)

    if "Признак МВЛ" in workbook.sheetnames:
        worksheet = workbook["Признак МВЛ"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            airport = normalize_key(row[0] if row else None)
            flag = as_float(row[1] if len(row) > 1 else None)
            if airport:
                international_airports[airport] = bool(flag)

    if "Справочники" in workbook.sheetnames:
        worksheet = workbook["Справочники"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            aircraft = normalize_key(row[5] if len(row) > 5 else None)
            multiplier = as_float(row[6] if len(row) > 6 else None)
            if aircraft and multiplier is not None:
                aircraft_multipliers[aircraft] = multiplier

            scenario = normalize_text(row[11] if len(row) > 11 else None)
            scenario_aircraft = normalize_key(row[12] if len(row) > 12 else None)
            rates = [as_float(row[index] if len(row) > index else None) for index in (13, 14, 15)]
            if scenario and scenario_aircraft and all(rate is not None for rate in rates):
                scenario_rates.setdefault(scenario, {})[scenario_aircraft] = [float(rate) for rate in rates if rate is not None]

    if "Прочее" in workbook.sheetnames:
        worksheet = workbook["Прочее"]
        # Строка 27 — значение «Итого» по аэропорту, которое использует формула
        # массива НО!F24. Это compatibility payload для one-time migration.
        for column in range(2, worksheet.max_column + 1):
            airport = normalize_key(worksheet.cell(1, column).value)
            amount = as_float(worksheet.cell(27, column).value)
            if airport and amount is not None:
                other_costs[airport] = amount

    # После обновления книги строки ЦРТ+ переходят в единый справочник веб-
    # приложения. Они остаются ручными и не удаляются обновлениями SRV.
    if "ЦРТ+" in workbook.sheetnames:
        worksheet = workbook["ЦРТ+"]
        iterator = worksheet.iter_rows(values_only=True)
        header = next(iterator, ())
        index = header_index(header)
        for row_number, row in enumerate(iterator, start=2):
            airport = normalize_key(value_by_key(row, index, "АП"))
            service = normalize_key(value_by_key(row, index, "Услуга"))
            rate = as_float(value_by_key(row, index, "Ставка"))
            if not airport or not service or rate is None:
                continue
            legacy_manual_tariffs.append(
                {
                    "id": f"legacy-{row_number}",
                    "airport": airport,
                    "service": service,
                    "rate": round(rate, 6),
                    "unit": normalize_text(value_by_key(row, index, "Ед.изм.")),
                    "aircraft": normalize_key(value_by_key(row, index, "В/С")),
                    "start_date": date_rank(value_by_key(row, index, "Дата с")),
                    "end_date": date_rank(value_by_key(row, index, "Дата по")),
                    "organization": normalize_text(value_by_key(row, index, "Наименование Орг")),
                    "note": "Перенесено из листа ЦРТ+ исходного монитора",
                    "source": "manual",
                    "source_file": path.name,
                    "source_row": row_number,
                    "legacy_manual": True,
                }
            )
    workbook.close()
    return {
        "routes": routes,
        "international_airports": international_airports,
        "other_costs": other_costs,
        "aircraft_multipliers": aircraft_multipliers,
        "scenario_rates": scenario_rates,
        "legacy_manual_tariffs": legacy_manual_tariffs,
    }, rows_read, preview, None
