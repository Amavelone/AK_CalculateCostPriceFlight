from __future__ import annotations

import re
import xml.etree.ElementTree as element_tree
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from ..catalog import normalize_key, normalize_text
from .common import as_float, header_index, normalize_for_json, value_by_key


@dataclass(frozen=True)
class ExchangeRateMetadata:
    rate: float
    source: str
    timestamp: str
    fallback_used: bool

    @property
    def note(self) -> str:
        return self.source if not self.fallback_used else "резервное значение 95 RUB/USD: ЦБ РФ недоступен"

    def __iter__(self):
        # Сохраняет backward-compatible распаковку для существующих вызовов adapter.
        yield self.rate
        yield self.note


def fetch_usd_rate() -> ExchangeRateMetadata:
    """Получает курс USD у ЦБ РФ и возвращает документированное резервное значение.

    Источник и запасной курс полностью повторяют действующее поведение Excel.
    При недоступности сети используется 95 RUB/USD с пояснением для интерфейса.
    """

    try:
        response = httpx.get("https://www.cbr.ru/scripts/XML_daily.asp", timeout=5.0)
        response.raise_for_status()
        document = element_tree.fromstring(response.content)
        for node in document.findall("Valute"):
            if node.findtext("CharCode") == "USD":
                rate = float((node.findtext("Value") or "").replace(",", "."))
                nominal = float((node.findtext("Nominal") or "1").replace(",", "."))
                return ExchangeRateMetadata(rate / nominal, "ЦБ РФ", datetime.now(UTC).isoformat(), False)
    except Exception:
        pass
    return ExchangeRateMetadata(95.0, "fallback", datetime.now(UTC).isoformat(), True)


def parse_fuel_registry(path: Path) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]], str | None]:
    """Извлекает канонические цены Fuel Registry и метаданные курса USD.

    Для каждого аэропорта сохраняется максимальная применимая цена. Конвертация
    USD использует ЦБ РФ с документированным fallback, который возвращается в
    заметке источника и сохраняется в provenance записи.
    """

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_row: tuple[Any, ...] | None = None
    iterator = worksheet.iter_rows(values_only=True)
    for row in iterator:
        if any(normalize_key(value) == "ПАРТНЕР" for value in row):
            header_row = row
            break
    if header_row is None:
        workbook.close()
        raise ValueError("Не найдена строка заголовков реестра (Партнер)")

    index = header_index(header_row)
    exchange_rate = fetch_usd_rate()
    if isinstance(exchange_rate, tuple):
        rate, source = exchange_rate
        exchange_rate = ExchangeRateMetadata(
            float(rate),
            str(source),
            datetime.now(UTC).isoformat(),
            "резерв" in str(source).lower() or "fallback" in str(source).lower(),
        )
    rows_read = 0
    preview: list[dict[str, Any]] = []
    fuel_by_airport: dict[str, dict[str, Any]] = {}

    for row in iterator:
        rows_read += 1
        price_kind = normalize_text(value_by_key(row, index, "Вид цены поставщика"))
        currency = normalize_key(value_by_key(row, index, "Валюта"))
        price = as_float(value_by_key(row, index, "Цена"))
        matches = re.findall(r"[A-Z]{3}", price_kind.upper())
        # В выгрузке 1С аэропорт содержится в виде цены, например
        # `-|MJZ|НДС сверху|RUB`; завершающая валюта не должна стать ключом.
        airports = [candidate for candidate in matches if candidate not in {"RUB", "USD", "EUR"}]
        airport = airports[0] if airports else ""
        if not airport or price is None:
            continue
        price_rub = price * exchange_rate.rate if currency == "USD" else price
        record = {
            "airport": airport,
            "price": round(price_rub, 6),
            "currency": currency or "RUB",
            "partner": normalize_text(value_by_key(row, index, "Партнер")),
            "price_kind": price_kind,
            "period": normalize_for_json(value_by_key(row, index, "Период")),
            "source_file": path.name,
        }
        if currency == "USD":
            record.update(
                {
                    "exchange_rate": exchange_rate.rate,
                    "exchange_rate_source": exchange_rate.source,
                    "exchange_rate_timestamp": exchange_rate.timestamp,
                    "exchange_rate_fallback_used": exchange_rate.fallback_used,
                }
            )
        current = fuel_by_airport.get(airport)
        # Утверждённое правило реестра сохраняет максимальную применимую цену
        # аэропорта; при доработке parser нельзя заменять его поиском первой строки.
        if current is None or record["price"] > current["price"]:
            fuel_by_airport[airport] = record
        if len(preview) < 12:
            preview.append(record)

    workbook.close()
    return list(fuel_by_airport.values()), rows_read, preview, exchange_rate.note
