from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..schemas import CalculationRequest


COMPONENT_LABELS = {
    "fuel": "ГСМ",
    "ground": "Наземное обслуживание",
    "ano": "Аэронавигационное обслуживание",
    "catering": "Бортовое питание",
    "vat": "НДС",
}
COMPONENT_ORDER = tuple(COMPONENT_LABELS)
CURRENCY_FORMAT = '#,##0.00;[Red](#,##0.00);-'
DECIMAL_FORMAT = '0.000;[Red](0.000);-'


def build_export_snapshot(request: CalculationRequest, result: dict[str, Any]) -> dict[str, Any]:
    """Create the single data model used by both JSON and XLSX exports.

    The function only packages the already calculated result. It deliberately
    contains no tariff lookups or business formulas, so export cannot alter a
    calculation.
    """

    inputs_by_id = {leg.id: leg.model_dump() for leg in request.legs}
    legs: list[dict[str, Any]] = []
    for number, result_leg in enumerate(result["legs"], start=1):
        input_leg = inputs_by_id.get(result_leg["id"], {})
        legs.append(
            {
                "number": number,
                "input": {
                    "id": result_leg["id"],
                    "departure": input_leg.get("departure", result_leg["departure"]),
                    "arrival": input_leg.get("arrival", result_leg["arrival"]),
                    "aircraft": input_leg.get("aircraft", result_leg["aircraft"]),
                    "passengers": input_leg.get("passengers", result_leg["passengers"]),
                },
                "result": {
                    "route": result_leg["route"],
                    "line_type": result_leg["line_type"],
                    "is_techstop": result_leg["is_techstop"],
                    "flight_time": result_leg["flight_time"],
                    "distance": result_leg["distance"],
                    "fuel_tons": result_leg["fuel_tons"],
                    "components": result_leg["components"],
                    "totals": result_leg["totals"],
                    "details": result_leg["details"],
                    "warnings": result_leg["warnings"],
                },
            }
        )

    return {
        "schema_version": "1.0",
        "exported_at": result["calculated_at"],
        "calculation": {
            "configuration": request.settings.model_dump(),
            "legs": legs,
            "totals": {
                "legs_count": len(legs),
                "flight_time": round(sum(leg["result"]["flight_time"] for leg in legs), 3),
                "fuel_tons": round(sum(leg["result"]["fuel_tons"] for leg in legs), 3),
                "m1": result["total"]["m1"],
                "m2": result["total"]["m2"],
                "m3": result["total"]["m3"],
            },
            "warnings": result["warnings"],
        },
    }


def json_bytes(snapshot: dict[str, Any]) -> bytes:
    return json.dumps(snapshot, ensure_ascii=False, indent=2).encode("utf-8")


def export_filename(snapshot: dict[str, Any], extension: str) -> str:
    exported_at = snapshot["exported_at"].replace("Z", "+00:00")
    timestamp = datetime.fromisoformat(exported_at).strftime("%Y-%m-%d_%H%M%S")
    return f"Расчет_себестоимости_{timestamp}.{extension}"


def xlsx_bytes(snapshot: dict[str, Any]) -> bytes:
    """Package the shared snapshot into an XLSX layout resembling РАСЧЕТ."""

    workbook = Workbook()
    calculation_sheet = workbook.active
    calculation_sheet.title = "РАСЧЕТ"
    details_sheet = workbook.create_sheet("ДЕТАЛИЗАЦИЯ")
    settings_sheet = workbook.create_sheet("ПАРАМЕТРЫ")
    _write_calculation_sheet(calculation_sheet, snapshot)
    _write_details_sheet(details_sheet, snapshot)
    _write_settings_sheet(settings_sheet, snapshot)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _header_style(cell: Any, fill: str, color: str = "FFFFFF") -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=10, bold=True, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_calculation_sheet(sheet: Any, snapshot: dict[str, Any]) -> None:
    headers = [
        "Номер плеча",
        "Ключ",
        "Маршрут",
        "Ставка НДС",
        "Flight Time",
        "ГСМ тонн",
        "ГСМ",
        "Наземное обслуживание",
        "Аэронавигационное обслуживание",
        "Бортовое питание",
        "НДС",
        "М1",
        "М2",
        "М3",
        "ИТОГО М1",
        "ИТОГО М2",
        "ИТОГО М3",
        "",
        "Тип ВС",
        "Аэропорт вылета IATA",
        "Аэропорт посадки IATA",
        "Загрузка, пасс.",
        "Вид линии",
        "Сценарий ЛЧ",
        "Стоимость ГСМ",
        "Техстоп",
        "Бортпитание",
    ]
    sheet.append(headers)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 42
    for index, cell in enumerate(sheet[1], start=1):
        _header_style(cell, "FFC000" if index >= 24 else "1F4E78", "1F1F1F" if index >= 24 else "FFFFFF")

    configuration = snapshot["calculation"]["configuration"]
    for leg in snapshot["calculation"]["legs"]:
        input_leg = leg["input"]
        result_leg = leg["result"]
        components = result_leg["components"]
        row = [
            leg["number"],
            f"{input_leg['aircraft']}-{input_leg['departure']}",
            result_leg["route"],
            0.1 if components["vat"] else 0,
            result_leg["flight_time"],
            result_leg["fuel_tons"],
            components["fuel"],
            components["ground"],
            components["ano"],
            components["catering"],
            components["vat"],
            components["m1"],
            components["m2"],
            components["m3"],
            result_leg["totals"]["m1"],
            result_leg["totals"]["m2"],
            result_leg["totals"]["m3"],
            None,
            input_leg["aircraft"],
            input_leg["departure"],
            input_leg["arrival"],
            input_leg["passengers"],
            result_leg["line_type"],
            configuration["scenario"],
            configuration["fuel_source"],
            "Да" if result_leg["is_techstop"] else "Нет",
            "Включено" if configuration["catering"] else "Не включено",
        ]
        sheet.append(row)

    totals = snapshot["calculation"]["totals"]
    sheet.append(
        [
            None,
            None,
            None,
            None,
            totals["flight_time"],
            totals["fuel_tons"],
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "ИТОГ",
            totals["m1"],
            totals["m2"],
            totals["m3"],
        ]
    )
    total_row = sheet.max_row
    for cell in sheet[total_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(name="Aptos", size=10, bold=True, color="173B59")
        cell.alignment = Alignment(vertical="center")

    money_columns = list(range(4, 18))
    for row_index in range(2, total_row + 1):
        sheet.cell(row_index, 4).number_format = "0.0%"
        sheet.cell(row_index, 5).number_format = DECIMAL_FORMAT
        sheet.cell(row_index, 6).number_format = DECIMAL_FORMAT
        for column in money_columns:
            if column not in (4, 5, 6):
                sheet.cell(row_index, column).number_format = CURRENCY_FORMAT
        for column in range(1, 28):
            cell = sheet.cell(row_index, column)
            cell.alignment = Alignment(vertical="center", horizontal="right" if 4 <= column <= 17 else "left")

    for row_index in range(2, total_row):
        for column in range(24, 28):
            sheet.cell(row_index, column).fill = PatternFill("solid", fgColor="FFF2CC")
    sheet.auto_filter.ref = f"A1:AA{total_row - 1}"
    widths = [12, 14, 14, 11, 12, 12, 15, 19, 23, 17, 14, 14, 14, 14, 17, 17, 17, 3, 12, 19, 20, 14, 12, 18, 15, 12, 15]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.column_dimensions["R"].width = 3
    thin = Side(style="thin", color="D9E2F3")
    for row in sheet.iter_rows(min_row=1, max_row=total_row, min_col=1, max_col=27):
        for cell in row:
            cell.border = Border(bottom=thin)


def _write_details_sheet(sheet: Any, snapshot: dict[str, Any]) -> None:
    headers = ["№ плеча", "Маршрут", "Компонент", "Аэропорт / маршрут", "Услуга / составляющая", "Ставка", "Объём", "Делитель", "Сумма"]
    sheet.append(headers)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        _header_style(cell, "1F4E78")

    for leg in snapshot["calculation"]["legs"]:
        result_leg = leg["result"]
        for component in COMPONENT_ORDER:
            for detail in result_leg["details"].get(component, []):
                sheet.append(
                    [
                        leg["number"],
                        result_leg["route"],
                        COMPONENT_LABELS[component],
                        detail.get("airport", ""),
                        detail["service"],
                        detail.get("rate"),
                        detail.get("volume"),
                        detail.get("divisor", 1),
                        detail["amount"],
                    ]
                )
    if sheet.max_row == 1:
        sheet.append([None, None, "Нет детализации", None, None, None, None, None, None])
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 6).number_format = CURRENCY_FORMAT
        sheet.cell(row_index, 7).number_format = DECIMAL_FORMAT
        sheet.cell(row_index, 8).number_format = DECIMAL_FORMAT
        sheet.cell(row_index, 9).number_format = CURRENCY_FORMAT
    sheet.auto_filter.ref = f"A1:I{sheet.max_row}"
    for column, width in enumerate([11, 15, 29, 20, 36, 15, 13, 11, 16], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_settings_sheet(sheet: Any, snapshot: dict[str, Any]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.append(["ПАРАМЕТР", "ЗНАЧЕНИЕ"])
    for cell in sheet[1]:
        _header_style(cell, "1F4E78")
    configuration = snapshot["calculation"]["configuration"]
    values = [
        ("Версия схемы", snapshot["schema_version"]),
        ("Дата и время расчёта (UTC)", snapshot["exported_at"]),
        ("Сценарий ЛЧ", configuration["scenario"]),
        ("Источник ГСМ", configuration["fuel_source"]),
        ("Техстоп", configuration["techstop_leg_id"] or "Не выбран"),
        ("Доплата за пассажиров", "Включено" if configuration["catering"] else "Не включено"),
        ("Детализация на экране", "Включено" if configuration["show_details"] else "Не включено"),
        ("Количество плеч", snapshot["calculation"]["totals"]["legs_count"]),
        ("Итоговый Flight Time, ч", snapshot["calculation"]["totals"]["flight_time"]),
        ("Итоговый ГСМ, тонн", snapshot["calculation"]["totals"]["fuel_tons"]),
        ("ИТОГО М1", snapshot["calculation"]["totals"]["m1"]),
        ("ИТОГО М2", snapshot["calculation"]["totals"]["m2"]),
        ("ИТОГО М3", snapshot["calculation"]["totals"]["m3"]),
    ]
    for value in values:
        sheet.append(list(value))
    warning_start = sheet.max_row + 2
    sheet.cell(warning_start, 1, "ПРЕДУПРЕЖДЕНИЯ")
    _header_style(sheet.cell(warning_start, 1), "A65D03")
    sheet.merge_cells(start_row=warning_start, start_column=1, end_row=warning_start, end_column=2)
    warnings = snapshot["calculation"]["warnings"] or ["Нет предупреждений"]
    for warning in warnings:
        sheet.append([warning, None])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=2)
        sheet.cell(sheet.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    for row_index in range(2, warning_start):
        if row_index in (10, 11):
            sheet.cell(row_index, 2).number_format = DECIMAL_FORMAT
        if row_index in (12, 13, 14):
            sheet.cell(row_index, 2).number_format = CURRENCY_FORMAT
    sheet.column_dimensions["A"].width = 33
    sheet.column_dimensions["B"].width = 58
