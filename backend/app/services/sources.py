from __future__ import annotations

import re
import shutil
import urllib.request
import uuid
import xml.etree.ElementTree as element_tree
from datetime import date, datetime, time
from itertools import chain
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


ALLOWED_SERVICES = {
    "АНО АД",
    "АЭРОВОКЗАЛ",
    "АЭРОВОКЗАЛ(М)",
    "БУКСИРОВКА",
    "ВЗЛЕТ-ПОСАДКА",
    "ВОДА",
    "ЗАПРАВКА ВС",
    "КЕРОСИН",
    "ОБЕСПЕЧЕНИЕ МАСЛАМИ ТН",
    "ПАССАЖИР",
    "ПАССАЖИР(М)",
    "ПРИЕМ-ВЫПУСК",
    "САНУЗЕЛ",
    "СЛИВ ВОДЫ",
    "ТЕЛЕТРАП МИН",
    "ТРАНСПБЕЗОП",
    "ТРАНСПОРТ",
    "ТРАП",
    "УБОРКА",
    "БОРТПИТАНИЕ",
}

# Exact AER selection rules from the active Power Query ЦРТ_Check definition.
AER_RULES: dict[str, str | None] = {
    "БУКСИРОВКА": None,
    "ВОДА": "738",
    "ПРИЕМ-ВЫПУСК": "738",
    "САНУЗЕЛ": "738",
    "БОРТПИТАНИЕ": "738",
    "УБОРКА": "737",
}


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_key(value: Any) -> str:
    return normalize_text(value).upper().replace("Ё", "Е")


def as_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (float, int)):
        return float(value)
    cleaned = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def as_hours(value: Any) -> float:
    if isinstance(value, time):
        # ИШР!P calculates HOUR(E) + MINUTE(E) / 60, intentionally ignoring
        # seconds even when the stored Excel time happens to contain them.
        return value.hour + value.minute / 60
    if isinstance(value, datetime):
        return value.hour + value.minute / 60
    number = as_float(value)
    if number is None:
        return 0.0
    return number * 24 if 0 <= number <= 1 else number


def date_rank(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return normalize_text(value)


def find_latest_file(source: dict[str, Any]) -> Path:
    directory = Path(source["directory"])
    if not directory.is_dir():
        raise FileNotFoundError(f"Директория не найдена: {directory}")
    candidates = [path for path in directory.glob(source["mask"]) if path.is_file()]
    if not candidates:
        raise FileNotFoundError(f"Не найден файл по маске {source['mask']}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def workbook_preview(path: Path, sheet_name: str | None = None, row_limit: int = 12) -> dict[str, Any]:
    """Return a readable preview of one workbook sheet.

    Registry exports place metadata before their actual headers, so the preview
    searches a small leading window rather than blindly treating row one as a
    table header. The UI can request any visible worksheet by name.
    """

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name and sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Лист не найден: {sheet_name}")
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    leading_rows: list[tuple[Any, ...]] = []
    for _ in range(25):
        row = next(iterator, None)
        if row is None:
            break
        leading_rows.append(row)

    header_position = 0
    header_hints = {"АП", "АЭРОПОРТ", "ПАРТНЕР", "УСЛУГА", "ВИД ЦЕНЫ ПОСТАВЩИКА"}
    for index, row in enumerate(leading_rows):
        values = {normalize_key(value) for value in row if normalize_text(value)}
        if values.intersection(header_hints):
            header_position = index
            break
    header = leading_rows[header_position] if leading_rows else ()
    labels = [normalize_text(value) or f"Колонка {index + 1}" for index, value in enumerate(header)]
    preview: list[dict[str, Any]] = []
    remaining_rows = leading_rows[header_position + 1 :]
    for row in chain(remaining_rows, iterator):
        if len(preview) >= row_limit:
            break
        preview.append({labels[index]: normalize_for_json(value) for index, value in enumerate(row) if index < len(labels)})
    active_sheet = worksheet.title
    sheet_names = list(workbook.sheetnames)
    workbook.close()
    return {"sheet": active_sheet, "sheets": sheet_names, "preview": preview}


def normalize_for_json(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def header_index(header: tuple[Any, ...]) -> dict[str, int]:
    return {normalize_key(value): index for index, value in enumerate(header) if normalize_text(value)}


def value_by_key(row: tuple[Any, ...], index: dict[str, int], key: str) -> Any:
    position = index.get(normalize_key(key))
    return row[position] if position is not None and position < len(row) else None


def parse_srv_tariffs(path: Path) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]], str | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    header = next(iterator, ())
    index = header_index(header)
    rows_read = 0
    candidates: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []

    for row_number, row in enumerate(iterator, start=2):
        rows_read += 1
        airport = normalize_key(value_by_key(row, index, "АП"))
        service = normalize_key(value_by_key(row, index, "Услуга"))
        aircraft = normalize_key(value_by_key(row, index, "В/С"))
        rate = as_float(value_by_key(row, index, "Ставка"))

        if len(preview) < 12:
            preview.append(
                {
                    "АП": airport,
                    "Услуга": service,
                    "Ставка": rate,
                    "В/С": aircraft,
                }
            )

        if not airport or service not in ALLOWED_SERVICES or rate is None:
            continue
        if aircraft not in {"", "733", "737", "738"}:
            continue
        candidates.append(
            {
                "id": f"srv-{row_number}",
                "airport": airport,
                "service": service,
                "rate": round(rate, 6),
                "unit": normalize_text(value_by_key(row, index, "Ед.изм.")),
                "aircraft": aircraft,
                "start_date": date_rank(value_by_key(row, index, "Дата с")),
                "end_date": date_rank(value_by_key(row, index, "Дата по")),
                "organization": normalize_text(value_by_key(row, index, "Наименование Орг")),
                "source": "file",
                "source_file": path.name,
                "source_row": row_number,
            }
        )
    workbook.close()

    # Current M-code leaves only one kerosene tariff per airport. For a tie the
    # deterministic ordering mirrors its stated aircraft/date tie-break intent.
    kerosene_by_airport: dict[str, dict[str, Any]] = {}
    for tariff in candidates:
        if tariff["service"] != "КЕРОСИН":
            continue
        current = kerosene_by_airport.get(tariff["airport"])
        comparison_key = (tariff["rate"], tariff["aircraft"], tariff["start_date"])
        if current is None or comparison_key > (current["rate"], current["aircraft"], current["start_date"]):
            kerosene_by_airport[tariff["airport"]] = tariff

    tariffs = [
        tariff
        for tariff in candidates
        if tariff["service"] != "КЕРОСИН" or kerosene_by_airport[tariff["airport"]]["id"] == tariff["id"]
    ]

    # Power Query preserves only maximum-rate AER rows for the six configured
    # services; five of them additionally require a particular aircraft type.
    aer_max_rates: dict[str, float] = {}
    for tariff in tariffs:
        if tariff["airport"] == "AER" and tariff["service"] in AER_RULES:
            aer_max_rates[tariff["service"]] = max(aer_max_rates.get(tariff["service"], tariff["rate"]), tariff["rate"])

    tariffs = [
        tariff
        for tariff in tariffs
        if tariff["airport"] != "AER"
        or tariff["service"] not in AER_RULES
        or (
            tariff["rate"] == aer_max_rates.get(tariff["service"])
            and (AER_RULES[tariff["service"]] is None or tariff["aircraft"] == AER_RULES[tariff["service"]])
        )
    ]
    return tariffs, rows_read, preview, None


def parse_nad_baseline(path: Path) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]], str | None]:
    """Preserve the active Excel baseline: NAD has no mapped `Ставка` column."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    header = next(iterator, ())
    index = header_index(header)
    rows_read = 0
    preview: list[dict[str, Any]] = []
    for row in iterator:
        rows_read += 1
        if len(preview) < 12:
            preview.append(
                {
                    "АП": normalize_key(value_by_key(row, index, "АП")),
                    "Услуга": normalize_key(value_by_key(row, index, "Услуга")),
                    "Надбавка/скидка": normalize_for_json(value_by_key(row, index, "Надбавка/скидка")),
                    "Ставка после M-кода": None,
                }
            )
    workbook.close()
    note = "Строки NAD прочитаны, но не добавлены: активный M-код не сопоставляет Надбавка/скидка со Ставка."
    return [], rows_read, preview, note


def fetch_usd_rate() -> tuple[float, str]:
    """Uses CBR as Excel does; returns a documented fallback if it is unavailable."""

    try:
        with urllib.request.urlopen("https://www.cbr.ru/scripts/XML_daily.asp", timeout=5) as response:
            document = element_tree.fromstring(response.read())
        for node in document.findall("Valute"):
            if node.findtext("CharCode") == "USD":
                rate = float((node.findtext("Value") or "").replace(",", "."))
                nominal = float((node.findtext("Nominal") or "1").replace(",", "."))
                return rate / nominal, "ЦБ РФ"
    except Exception:
        pass
    return 95.0, "резервное значение 95 RUB/USD: ЦБ РФ недоступен"


def parse_fuel_registry(path: Path) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]], str | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_row: tuple[Any, ...] | None = None
    iterator = worksheet.iter_rows(values_only=True)
    for row in iterator:
        if any(normalize_key(value) == "ПАРТНЕР" for value in row):
            header_row = row
            break
    if header_row is None:
        workbook.close()
        raise ValueError("Не найдена строка заголовков реестра (Партнер)")

    index = header_index(header_row)
    usd_rate, rate_note = fetch_usd_rate()
    rows_read = 0
    preview: list[dict[str, Any]] = []
    fuel_by_airport: dict[str, dict[str, Any]] = {}

    for row in iterator:
        rows_read += 1
        price_kind = normalize_text(value_by_key(row, index, "Вид цены поставщика"))
        currency = normalize_key(value_by_key(row, index, "Валюта"))
        price = as_float(value_by_key(row, index, "Цена"))
        matches = re.findall(r"[A-Z]{3}", price_kind.upper())
        # In the 1C export the airport appears in the price type, e.g.
        # `-|MJZ|НДС сверху|RUB`; the trailing currency must not become the key.
        airports = [candidate for candidate in matches if candidate not in {"RUB", "USD", "EUR"}]
        airport = airports[0] if airports else ""
        if not airport or price is None:
            continue
        price_rub = price * usd_rate if currency == "USD" else price
        record = {
            "airport": airport,
            "price": round(price_rub, 6),
            "currency": currency or "RUB",
            "partner": normalize_text(value_by_key(row, index, "Партнер")),
            "price_kind": price_kind,
            "period": normalize_for_json(value_by_key(row, index, "Период")),
            "source_file": path.name,
        }
        current = fuel_by_airport.get(airport)
        if current is None or record["price"] > current["price"]:
            fuel_by_airport[airport] = record
        if len(preview) < 12:
            preview.append(record)

    workbook.close()
    return list(fuel_by_airport.values()), rows_read, preview, rate_note


def parse_monitor_workbook(path: Path) -> tuple[dict[str, Any], int, list[dict[str, Any]], str | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    routes: list[dict[str, Any]] = []
    international_airports: dict[str, bool] = {}
    other_costs: dict[str, float] = {}
    aircraft_multipliers: dict[str, float] = {}
    scenario_rates: dict[str, dict[str, list[float]]] = {}
    legacy_manual_tariffs: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []
    rows_read = 0

    if "ИШР" in workbook.sheetnames:
        worksheet = workbook["ИШР"]
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            rows_read += 1
            departure = normalize_key(row[1] if len(row) > 1 else None)
            arrival = normalize_key(row[2] if len(row) > 2 else None)
            distance = as_float(row[3] if len(row) > 3 else None) or 0.0
            flight_time = as_hours(row[4] if len(row) > 4 else None)
            if not departure or not arrival:
                continue
            route = {
                "key": f"{departure}-{arrival}",
                "departure": departure,
                "arrival": arrival,
                "distance": round(distance, 3),
                # Power Query preserves the full fractional Excel time. Do not round
                # it to display precision: a few seconds affect fuel and route totals.
                "flight_time": round(flight_time, 12),
                "source_row": row_number,
            }
            routes.append(route)
            if len(preview) < 12:
                preview.append(route)

    if "Признак МВЛ" in workbook.sheetnames:
        worksheet = workbook["Признак МВЛ"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            airport = normalize_key(row[0] if row else None)
            flag = as_float(row[1] if len(row) > 1 else None)
            if airport:
                international_airports[airport] = bool(flag)

    if "Справочники" in workbook.sheetnames:
        worksheet = workbook["Справочники"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            aircraft = normalize_key(row[5] if len(row) > 5 else None)
            multiplier = as_float(row[6] if len(row) > 6 else None)
            if aircraft and multiplier is not None:
                aircraft_multipliers[aircraft] = multiplier

            scenario = normalize_text(row[11] if len(row) > 11 else None)
            scenario_aircraft = normalize_key(row[12] if len(row) > 12 else None)
            rates = [as_float(row[index] if len(row) > index else None) for index in (13, 14, 15)]
            if scenario and scenario_aircraft and all(rate is not None for rate in rates):
                scenario_rates.setdefault(scenario, {})[scenario_aircraft] = [float(rate) for rate in rates if rate is not None]

    if "Прочее" in workbook.sheetnames:
        worksheet = workbook["Прочее"]
        # Row 27 is the per-airport "Итого" value used by the array formula
        # in НО!F24. Keep it as configuration rather than a hard-coded rule.
        for column in range(2, worksheet.max_column + 1):
            airport = normalize_key(worksheet.cell(1, column).value)
            amount = as_float(worksheet.cell(27, column).value)
            if airport and amount is not None:
                other_costs[airport] = amount

    # Existing ЦРТ+ rows are migrated into the unified web directory once the
    # workbook is refreshed. They remain manual rows and are never deleted by
    # SRV/NAD updates.
    if "ЦРТ+" in workbook.sheetnames:
        worksheet = workbook["ЦРТ+"]
        iterator = worksheet.iter_rows(values_only=True)
        header = next(iterator, ())
        index = header_index(header)
        for row_number, row in enumerate(iterator, start=2):
            airport = normalize_key(value_by_key(row, index, "АП"))
            service = normalize_key(value_by_key(row, index, "Услуга"))
            rate = as_float(value_by_key(row, index, "Ставка"))
            if not airport or not service or rate is None:
                continue
            legacy_manual_tariffs.append(
                {
                    "id": f"legacy-{row_number}",
                    "airport": airport,
                    "service": service,
                    "rate": round(rate, 6),
                    "unit": normalize_text(value_by_key(row, index, "Ед.изм.")),
                    "aircraft": normalize_key(value_by_key(row, index, "В/С")),
                    "start_date": date_rank(value_by_key(row, index, "Дата с")),
                    "end_date": date_rank(value_by_key(row, index, "Дата по")),
                    "organization": normalize_text(value_by_key(row, index, "Наименование Орг")),
                    "note": "Перенесено из листа ЦРТ+ исходного монитора",
                    "source": "manual",
                    "source_file": path.name,
                    "source_row": row_number,
                    "legacy_manual": True,
                }
            )
    workbook.close()
    return {
        "routes": routes,
        "international_airports": international_airports,
        "other_costs": other_costs,
        "aircraft_multipliers": aircraft_multipliers,
        "scenario_rates": scenario_rates,
        "legacy_manual_tariffs": legacy_manual_tariffs,
    }, rows_read, preview, None


PARSERS: dict[str, Callable[[Path], tuple[Any, int, list[dict[str, Any]], str | None]]] = {
    "srv_tariffs": parse_srv_tariffs,
    "nad_baseline": parse_nad_baseline,
    "fuel_registry": parse_fuel_registry,
    "monitor_workbook": parse_monitor_workbook,
}


def source_by_id(state: dict[str, Any], source_id: str) -> dict[str, Any]:
    for source in state["source_configs"]:
        if source["id"] == source_id:
            return source
    raise KeyError(source_id)


def refresh_source(state: dict[str, Any], source_id: str, now: str) -> dict[str, Any]:
    source = source_by_id(state, source_id)
    path = find_latest_file(source)
    parser = PARSERS[source["parser"]]
    result, rows_read, preview, note = parser(path)

    if source_id == "srv":
        state["imported_tariffs"] = result
        rows_loaded = len(result)
    elif source_id == "fuel_registry":
        state["fuel_prices"] = result
        rows_loaded = len(result)
    elif source_id == "monitor_workbook":
        state["routes"] = result["routes"]
        state["international_airports"] = result["international_airports"]
        state["other_costs"] = result["other_costs"]
        if result["aircraft_multipliers"]:
            state["aircraft_multipliers"] = result["aircraft_multipliers"]
        if result["scenario_rates"]:
            state["scenario_rates"] = result["scenario_rates"]
        non_legacy_manual = [item for item in state["manual_tariffs"] if not item.get("legacy_manual")]
        state["manual_tariffs"] = non_legacy_manual + result["legacy_manual_tariffs"]
        rows_loaded = len(result["routes"])
    else:
        rows_loaded = len(result)

    source.update(
        {
            "last_status": "ready",
            "last_file": path.name,
            "last_updated": now,
            "last_error": None,
            "last_note": note,
            "rows_read": rows_read,
            "rows_loaded": rows_loaded,
            "preview": preview,
        }
    )
    return source


def mark_source_error(state: dict[str, Any], source_id: str, message: str, now: str) -> dict[str, Any]:
    source = source_by_id(state, source_id)
    source.update({"last_status": "error", "last_error": message, "last_updated": now})
    return source


def save_uploaded_file(source: dict[str, Any], original_name: str, source_file: Any) -> Path:
    file_name = Path(original_name or "source.xlsx").name
    if Path(file_name).suffix.lower() != ".xlsx":
        raise ValueError("Поддерживаются только файлы .xlsx")
    directory = Path(source["directory"])
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / file_name
    temporary = directory / f".{file_name}.{uuid.uuid4().hex}.uploading"
    with temporary.open("wb") as target_file:
        shutil.copyfileobj(source_file, target_file)
    temporary.replace(target)
    return target


def tariffs_for_view(state: dict[str, Any]) -> list[dict[str, Any]]:
    imported = [dict(item, conflict=False) for item in state.get("imported_tariffs", [])]
    imported_keys = {f"{item['airport']}-{item['service']}" for item in imported}
    manual = [
        dict(item, conflict=f"{item['airport']}-{item['service']}" in imported_keys)
        for item in state.get("manual_tariffs", [])
    ]
    return imported + manual
